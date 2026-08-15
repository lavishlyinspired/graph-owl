"""Report exporters: CSV working paper, Excel working paper, ITC register."""

from __future__ import annotations

import csv
import io

from . import reconciliation as rc


def _flat_rows(results: list[dict]) -> list[dict]:
    rows = []
    for row in results:
        book = row["book"]
        portal = row["portal"]
        status = row["status"]
        rows.append(
            {
                "Status": {
                    rc.STATUS_MATCHED: "Matched",
                    rc.STATUS_REVIEW: "Review",
                    rc.STATUS_ONLY_BOOKS: "Only Books",
                    rc.STATUS_ONLY_GSTR2B: "Only GSTR-2B",
                }[status],
                "GSTIN": (book or portal)["gstin"],
                "Supplier": (book or portal)["supplier"],
                "Inv (GSTR-2B)": portal["invoice_no"] if portal else "—",
                "Inv (Books)": book["invoice_no"] if book else "—",
                "Voucher No": (book or {}).get("voucher_no", "") or "—",
                "Taxable (GSTR-2B)": portal["taxable"] if portal else None,
                "Tax (GSTR-2B)": portal["tax"] if portal else None,
                "Taxable (Books)": book["taxable"] if book else None,
                "Tax (Books)": book["tax"] if book else None,
                "Diff": row["diff"],
                "ITC Amt": row["itc"],
                "Reason": row["reason"],
                "Conf": _confidence(status),
            }
        )
    return rows


def _confidence(status: str) -> str:
    if status == rc.STATUS_MATCHED:
        return "100% OK"
    if status == rc.STATUS_REVIEW:
        return "Partial Match"
    return "—"


def _to_float(value) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def export_working_paper_csv(results: list[dict]) -> bytes:
    rows = _flat_rows(results)
    buffer = io.StringIO()
    fieldnames = list(rows[0].keys()) if rows else []
    writer = csv.DictWriter(buffer, fieldnames=fieldnames)
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    return buffer.getvalue().encode("utf-8")


def _xlsx_out(rows: list[dict], sheet_name: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_fill = PatternFill(start_color="1A2E1F", end_color="1A2E1F", fill_type="solid")
    header_font = Font(color="4ADE80", bold=True)
    num_fmt = "#,##0.00"

    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in rows:
            ws.append([row.get(h) for h in headers])

        numeric_cols = {
            h for h in headers if any(k in h.lower() for k in ("taxable", "tax", "diff", "amt"))
        }
        for col_idx, header in enumerate(headers, start=1):
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if header in numeric_cols:
                    value = _to_float(cell.value)
                    cell.value = value if value else None
                    cell.number_format = num_fmt
                cell.alignment = Alignment(vertical="center")

        widths = {h: max(len(h), 12) for h in headers}
        for row in rows:
            for h in headers:
                value = row.get(h)
                if value is not None:
                    widths[h] = max(widths[h], min(len(str(value)), 40))
        for col_idx, header in enumerate(headers, start=1):
            ws.column_dimensions[chr(64 + col_idx)].width = widths[header] + 2

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def export_working_paper_xlsx(results: list[dict]) -> bytes:
    return _xlsx_out(_flat_rows(results), "Working Paper")


def export_itc_register(results: list[dict]) -> bytes:
    rows = []
    for row in results:
        if row["status"] == rc.STATUS_ONLY_BOOKS:
            continue
        book = row["book"]
        portal = row["portal"]
        rows.append(
            {
                "GSTIN": (book or portal)["gstin"],
                "Supplier": (book or portal)["supplier"],
                "Invoice No": (book or portal)["invoice_no"],
                "Invoice Date": (book or portal).get("invoice_date", ""),
                "Taxable": (book or portal)["taxable"],
                "ITC Claimed (Tax)": row["itc"],
                "Status": {
                    rc.STATUS_MATCHED: "Matched",
                    rc.STATUS_REVIEW: "Review",
                    rc.STATUS_ONLY_GSTR2B: "Only GSTR-2B",
                }[row["status"]],
            }
        )
    return _xlsx_out(rows, "ITC Register")
